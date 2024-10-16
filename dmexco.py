from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import pickle
import time
import random
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import ElementNotInteractableException
import openpyxl
import re

# How to use the selenium template
# 1. Run cmd and type python -m venv venv
# 2. Activate the virtual environment by running venv\Scripts\activate
# 3. pip install -r requirements.txt AND pip install openpyxl
# 4. Change the USER_DATA_DIRECTORY and YOUR_USER_AGENT variables
# 5. Use the functions below to interact with the browser
# 6. Run the script


# get element by xpath
# driver.find_element(By.XPATH, xpath)


# get element by class name
# driver.find_element(By.CLASS_NAME, class_name)


# get element by id
# driver.find_element(By.ID, id)


# switch to iframe
# driver.switch_to.frame(iframe)


# switch to default content
# driver.switch_to.default_content()

# get the user data directory by running chrome://version/ in the browser and change the path accordingly
USER_DATA_DIRECTORY = r"C:\Users\areed\AppData\Local\Google\Chrome\User Data\Default"

# how to get user agent: https://www.whatismybrowser.com/detect/what-is-my-user-agent
YOUR_USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36"


def save_cookies(driver):
    pickle.dump(driver.get_cookies(), open("cookies.pkl", "wb"))

def load_cookies(driver):
    try:
        cookies = pickle.load(open("cookies.pkl", "rb"))
        for cookie in cookies:
            driver.add_cookie(cookie)
    except:
        pass

def browser_init():
    chrome_options = Options()
    chrome_options.add_argument("user-agent=" + YOUR_USER_AGENT)
    chrome_options.add_argument("--user-data-dir=" + USER_DATA_DIRECTORY)
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    driver.get('https://www.google.com/')
    load_cookies(driver)
    return driver


def random_sleep():
    time.sleep(random.randint(1, 4))

def wait_for_element(driver, xpath, timeout=30):
    start_time = time.time()
    while True:
        try:
            element = driver.find_element(By.XPATH, xpath)
            return element
        except:
            current_time = time.time()
            elapsed_time = current_time - start_time
            if elapsed_time > timeout:
                print(f"Timeout reached after {timeout} seconds")
                break
            pass
        


target = "https://community.dmexco.com/widget/event/dmexco-2024/exhibitors/RXZlbnRWaWV3XzgyMTMyMg==?showActions=true"
company_list_xpath = "/html/body/div[2]/main/div/div[2]/div/div/div/div/a"




def get_company_list(driver):
    driver.get(target)
    time.sleep(5)
    # scroll to the bottom of the page that has an infinite scroll
    for i in range(0, 20):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        random_sleep()
    
    company_list = driver.find_elements(By.XPATH, company_list_xpath)
    company_links = []
    for company in company_list:
        company_links.append(company.get_attribute("href"))
    
    return company_links
        

def get_links_from_txt():
    with open("company_links.txt", "r") as text_document:
        company_links = text_document.readlines()
    return company_links

def print_links_to_txt(company_links):
    with open("company_links.txt", "w") as text_document:
        for link in company_links:
            text_document.write(link + "\n")



def grab_text_info(driver, xpath_item):
    try:
        item = driver.find_element(By.XPATH, xpath_item).text
    except:
        item = "Not Found"
    return item



def extract_company_info(driver, company_links):
    for link in company_links:
        driver.get(link)
        time.sleep(5)
        company_name = grab_text_info(driver, '//*[contains(@class, "dbnofQ")]')
        sponsorship = grab_text_info(driver, '//*[contains(@class, "kgsBis")]')
        tags_list = driver.find_elements(By.XPATH, '//*[contains(@class, "dDzhoy")]/span')
        tags = []
        for tag in tags_list:
            tags.append(tag.text)
        

        description = []
        description_list = driver.find_elements(By.XPATH, '//*[contains(@class, "gwATgs")]/p')
        for desc in description_list:
            description.append(desc.text)
        
        socials = driver.find_elements(By.XPATH, '//*[contains(@class, "kzrhIj")]/a')

        social_links = []
        for social in socials:
            social_links.append(social.get_attribute("href"))

        contact_info = []

        contact_details = driver.find_elements(By.XPATH, '//*[contains(@class, "eQHIVq")]')
        for contact in contact_details:
            contact_info.append(contact.text)
        

        if len(driver.find_elements(By.XPATH, '//*[contains(@class, "djxlpz")]/span[contains(text(), "See all")]')) > 0:
            try:
                driver.find_element(By.XPATH, '//*[contains(@class, "djxlpz")]/span[contains(text(), "See all")]').click()
            except:
                pass
            time.sleep(2)




        team_members = driver.find_elements(By.XPATH, '//*[contains(@class, "livcmb")]/a')
        team_member_links = []

        for member in team_members:
            team_member_links.append(member.get_attribute("href"))
        export_to_excel(company_name, sponsorship, tags, description, social_links, contact_info, team_member_links, link)
        print(company_name + " extracted")

    
def export_to_excel(company_name, sponsorship, tags, description, social_links, contact_info, team_member_links, link):
    # open a workbook called Company Info.xlsx
    workbook = openpyxl.load_workbook("Company Info.xlsx")
    sheet = workbook.active
    tags = ", ".join(tags)
    description = "; ".join(description)
    social_links = ", ".join(social_links)
    contact_info = ", ".join(contact_info)
    team_member_links = ", ".join(team_member_links)

    # append the entire row
    sheet.append([company_name, sponsorship, tags, description, social_links, contact_info, team_member_links, link])

    # save the workbook
    workbook.save("Company Info.xlsx")


def create_company_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Company Name"
    sheet["B1"] = "Sponsorship"
    sheet["C1"] = "Tags"
    sheet["D1"] = "Description"
    sheet["E1"] = "Social Links"
    sheet["F1"] = "Contact Info"
    sheet["G1"] = "Team Member Links"
    sheet["H1"] = "Event Link"

    workbook.save("Company Info.xlsx")

def create_prospect_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Company Name"
    sheet["B1"] = "Contact Name"
    sheet["C1"] = "Contact Title"
    sheet["D1"] = "About"
    sheet["E1"] = "Country"
    sheet["F1"] = "Website"
    sheet["G1"] = "Social Links"
    sheet["H1"] = "Contact Info"
    sheet["I1"] = "Going to"
    sheet["J1"] = "Is Speaking"
    sheet["K1"] = "Sponsorship"
    sheet["L1"] = "Tags"
    sheet["M1"] = "Description"
    sheet["N1"] = "Company Social Links"
    sheet["O1"] = "Event Link"
    sheet["P1"] = "Prospect Link"

    workbook.save("Prospect Info.xlsx")

def export_prospect_to_excel(company_name, contact_name, contact_title, about, country, company_link, contact_social_links, contact_info, going_to, is_speaking, sponsorship, tags, description, company_socials, event_link, link):
    workbook = openpyxl.load_workbook("Prospect Info.xlsx")
    sheet = workbook.active
    sheet.append([company_name, contact_name, contact_title, about, country, company_link, contact_social_links, contact_info, going_to, is_speaking, sponsorship, tags, description, company_socials, event_link, link])
    workbook.save("Prospect Info.xlsx")


def scrape_prospect(driver, link):
    driver.get(link)
    time.sleep(1)
    random_sleep()

    contact_name = grab_text_info(driver, '//*[contains(@class, "hbvnEe")]')
    contact_title = grab_text_info(driver, '//*[contains(@class, "gqcPin")]')

    about = grab_text_info(driver, '//*[contains(@class, "hwZDkN")]')
    country = grab_text_info(driver, '//*[contains(@class, "hJCvjG") and text()="Country"]/following-sibling::div[1]')
    contact_socials = driver.find_elements(By.XPATH, '//*[contains(@class, "kzrhIj")]/a')
    social_links = []
    for social in contact_socials:
        social_links.append(social.get_attribute("href"))

    contact_social_links = ", ".join(social_links)

    contact_info = []
    contact_details = driver.find_elements(By.XPATH, '//*[contains(@class, "eQHIVq")]')
    for contact in contact_details:
        contact_info.append(contact.text)

    contact_info = ", ".join(contact_info)

    going_to = grab_text_info(driver, '//*[contains(@class, "gyWXZ")]')
    is_speaking_text = grab_text_info(driver, '//*[contains(@class, "jzoDyF")]')

    is_speaking = False

    if is_speaking_text == "Is speaking at":
        is_speaking = True
    
    full_contact_info = {}

    full_contact_info["contact_name"] = contact_name
    full_contact_info["contact_title"] = contact_title
    full_contact_info["about"] = about
    full_contact_info["country"] = country
    full_contact_info["contact_social_links"] = contact_social_links
    full_contact_info["contact_info"] = contact_info
    full_contact_info["going_to"] = going_to
    full_contact_info["is_speaking"] = is_speaking

    return full_contact_info

    
        
def read_company_info():
    workbook = openpyxl.load_workbook("Company Info.xlsx")
    sheet = workbook.active
    company_info = {}
    for row in sheet.iter_rows(values_only=True):
        # skip header
        if row[0] == "Company Name":
            continue
        company_info[row[0]] = {
            "company_name": row[0],
            "sponsorship": row[1],
            "tags": row[2],
            "description": row[3],
            "social_links": row[4],
            "contact_info": row[5],
            "team_member_links": row[6],
            "event_link": row[7]
        }
    return company_info

def start_scrape_prospects(company_info):
    for key in company_info:
        prospect_links = company_info[key]["team_member_links"].split(", ")

        url_pattern = r'(https?://[^\s]+)'
        try:
            company_link = re.findall(url_pattern, company_info[key]["contact_info"])[0]
        except:
            company_link = "Not Found"

        for link in prospect_links:
            full_contact_info = scrape_prospect(driver, link)
            export_prospect_to_excel(company_info[key]["company_name"], full_contact_info["contact_name"], full_contact_info["contact_title"], full_contact_info["about"], full_contact_info["country"], company_link, full_contact_info["contact_social_links"], full_contact_info["contact_info"], full_contact_info["going_to"], full_contact_info["is_speaking"], company_info[key]["sponsorship"], company_info[key]["tags"], company_info[key]["description"], company_info[key]["social_links"], company_info[key]["event_link"], link)



    





if __name__ == "__main__":
    driver = browser_init()
    # company_links = get_company_list(driver)
    # print_links_to_txt(company_links)
    create_company_excel()
    create_prospect_excel()
    
    company_links = get_links_from_txt()
    extract_company_info(driver, company_links)
    company_info = read_company_info()
    start_scrape_prospects(company_info)
    driver.quit()
    

    


    



